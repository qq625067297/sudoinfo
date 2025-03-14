from openpyxl import Workbook, load_workbook
import os
import re

wb = load_workbook(r"Yundu_回片系统验证_03.xlsx")



"""
       Number of Tests SKIPPED:       0
       Number of Tests PASSED:        0
       Number of Tests FAILED:        3
 Total Number of Tests RUN:           3
"""

def get_result(id, usp, mt):
    """
    :param id: cv id, TD_xx_xx
    :param usp: True: USP, False: DSP
    :param mt: True:众星微， False:数渡
    :return:
    """
    info = None
    SWITCH = ['数渡', '众星微'][mt]
    DEVICETYPE = ['DSP', 'USP'][usp]
    path = f'{SWITCH}/{DEVICETYPE}/'
    for filename in os.listdir(path):
        if id in filename:
            with open(os.path.join(path, filename), 'rb') as f:
                content = f.read().decode()
                # print(content)
                info = re.findall(rf'Number of Tests SKIPPED:\s+(\d+).*?Number of Tests PASSED:\s+(\d+).*?Number of Tests FAILED:\s+(\d+)', content, re.DOTALL)
                if info[0][0] != '0':
                    status = 'SKIPPED'
                elif info[0][2] != '0':
                    status = 'FAIL'
                else:
                    status = 'PASS'
                return status


ws = wb['PCIECV_USP']
for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
    if row[3].value:
        id = re.findall(r"(TD_\d+_\d+)", row[3].value)
        if id:
            result = get_result(id[0], True, True)
            ws.cell(row=row_idx, column=10).value = result

ws = wb['PCIECV_DSP']
for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
    if row[3].value:
        id = re.findall(r"(TD_\d+_\d+)", row[3].value)
        if id:
            result = get_result(id[0], False, True)
            ws.cell(row=row_idx, column=10).value = result

wb.save("1.xlsx")